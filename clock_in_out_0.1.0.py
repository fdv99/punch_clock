# python 3.7
# Clock in and out and save times to excel sheet
# 0.1.0 changed file saves to f strings

import datetime
import openpyxl
import string
import calendar
import os.path
import os

users = {
    "harisong": "greer",
    "chasev": "mcv",
    "blaked": "batman",
    "forrest": "fdv",
}

file_header = ["Date", "Weekday", "Clock In Time", "Clock Out Time", "Hours Worked", "Weekly Total", "Weekly Overtime"]

def clockInSheet(user, date, log_time, week_day, month, year, next_month, last_previous_month):
    # Return wb.active, take file path as argument
    if int(day) > 25: # If it is December, need to save as key_nextYear.01.xlsx
        year_str = next_month.strftime("%Y")
        month_str = next_month.strftime("%m")
        end_month = int(last_this_month.strftime("%d"))    
    else:    
        year_str = str(year)
        month_str = str(month)
        end_month = int(last_previous_month.strftime("%d"))
       
    if os.path.isfile(f"{user}_{year_str}.{month_str}.xlsx") == True:
        print("The file exists")
    else: 
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in string.ascii_uppercase:
            ws.column_dimensions[i].width = 14
        # Get number of days in given month, and enter all dates in first column, starting at row 3
        # Account for months with 31, 30, 29, and 28 days.
        for x in range(26, end_month + 1): 
            ws.cell(row=x - 23, column=1, value= f"last_previous_month.strftime("%Y").last_previous_month.strftime("%m").x)
        for y in range(1, 26):
            if end_month == 28:
                ws.cell(row=y + 5, column=1, value= str(year) + "." + str(month) + "." + str(y))
            if end_month == 29:
                ws.cell(row=y + 6, column=1, value= str(year) + "." + str(month) + "." + str(y))
            if end_month == 30:
                ws.cell(row=y + 7, column=1, value= str(year) + "." + str(month) + "." + str(y))
            if end_month == 31:
                ws.cell(row=y + 8, column=1, value= str(year) + "." + str(month) + "." + str(y))
        ws.cell(row=1, column=1, value= user)
        for j in range(1, len(file_header)): # Adding the column headers
            ws.cell(row=2, column=j, value= file_header[j])
        wb.save(f"{user}_{year_str}.{month_str}.xlsx")

    wb = openpyxl.load_workbook(f"{user}_{year_str}.{month_str}.xlsx")
    ws = wb.active
    for num1 in range(3, 35):
        if ws.cell(row=num1, column=1).value == date:
            ws.cell(row=num1, column=3, value= log_time)
            ws.cell(row=num1, column=2, value= week_day)
    wb.save(user + "_" + year_str + "." + month_str + ".xlsx")
    wa = openpyxl.load_workbook("userstatus.xlsx")
    wt = wa.active
    for num2 in range(1,12):
        if wt.cell(row=num2, column=1).value == user:
            wt.cell(row=num2, column=2, value= "in")
    wa.save("userstatus.xlsx")
    

def clockOutSheet(user, date, log_time, week_day, year, next_month):
    # arguments are row, log_time, week_day 
    if int(day) > 25: # If day is greater than 25th, need to move to next document
        year_str = next_month.strftime("%Y")
        month_str = next_month.strftime("%m")
    else: # On days 1-25    
        year_str = str(year)
        month_str = str(month)
    wb = openpyxl.load_workbook(f"{user}_{year_str}.{month_str}.xlsx")
    ws = wb.active
    for num in (range(3, 35)):
        if ws.cell(row=num, column=1).value == date:
            ws.cell(row=num, column=4, value= log_time)
            in_time = ws.cell(row=num, column=3).value
            out_time = ws.cell(row=num, column=4).value
            ws.cell(row=num, column=5, value= out_time - in_time)
            ws["E34"] = "=SUM(E3: E33)"
    wb.save(f"{user}_{year_str}.{month_str}.xlsx")
    wa = openpyxl.load_workbook("userstatus.xlsx")
    wt = wa.active
    for num2 in range(1,12):
        if wt.cell(row=num2, column=1).value == user:
            wt.cell(row=num2, column=2, value= "out")
    wa.save("userstatus.xlsx")

while True:
    active_user = input("User: q to quit ").lower().strip()

    if active_user in users:
        password = input("Enter password: ")
    elif active_user == "q":
        break
    else:
        print("That is not a valid user!")
        continue
    
    if password == users[active_user]:
        time = datetime.datetime.now()
        day = datetime.datetime.today().strftime("%d")
        date = (time.strftime("%Y.%#m.%#d"))
        print(time)
        log_time = float(time.strftime("%H")) + (float(time.strftime("%M"))/60)
        month = int(datetime.datetime.today().strftime("%m"))
        year = int(datetime.datetime.today().strftime("%Y"))
        week_day = datetime.datetime.today().strftime("%A")
        last_previous_month = time.replace(day=1) - datetime.timedelta(days=1)
        previous_month = last_previous_month.strftime("%m")
        next_month = time.replace(day=28) + datetime.timedelta(days=4)
        last_this_month = next_month.replace(day=1) - datetime.timedelta(days=1)

    else:
        print("Wrong Password!")
        continue

    wa = openpyxl.load_workbook("userstatus.xlsx")
    wt = wa.active
    for num2 in range(1,12):
        if wt.cell(row=num2, column=1).value == active_user:
            user_status = wt.cell(row=num2, column=2).value
    wa.save("userstatus.xlsx")

    # Check if user is clocked in or out
    if user_status == "out":
        print("You are now Clocked In.")
        # Add log_time to column 2 in user spreadsheet
        clockInSheet(active_user, date, log_time, week_day, month, year, next_month, last_previous_month)
        question = input("Any key to continue: ")
        os.system('cls')
        continue
        
    else:
        print("You are now Clocked Out.")
        # Add log_time to column 3 in user spreadsheet
        clockOutSheet(active_user, date, log_time, week_day, year, next_month)
        question2 = input("Any key to continue: ")
        os.system('cls')
        continue